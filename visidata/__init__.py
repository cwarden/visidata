#!/usr/bin/env python3
#
# VisiData: a curses interface for exploring and arranging tabular data
#
# Copyright (C) 2016 Saul Pwanson
#
#   This program is free software: you can redistribute it and/or modify
#   it under the terms of the GNU General Public License as published by
#   the Free Software Foundation, either version 3 of the License, or
#   (at your option) any later version.
#
#   This program is distributed in the hope that it will be useful,
#   but WITHOUT ANY WARRANTY; without even the implied warranty of
#   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#   GNU General Public License for more details.
#
#   You should have received a copy of the GNU General Public License
#   along with this program.  If not, see <http://www.gnu.org/licenses/>.

__version__ = '0.39'
__author__ = 'Saul Pwanson <vd@saul.pw>'
__license__ = 'GPLv3'
__status__ = 'Development'

import itertools
import os.path
from copy import copy
import io
import datetime
import string
import collections
import functools
import statistics
import curses
import curses.ascii

import re
import csv

from visidata.tui import edit_text, Key, Shift, Ctrl, keyname, EscapeException


initialStatus = 'saul.pw/VisiData v' + __version__
defaultColNames = list(itertools.chain(string.ascii_uppercase, [''.join(i) for i in itertools.product(string.ascii_uppercase, repeat=2)]))

g_vd = None  # toplevel VisiData, contains all sheets
scr = None   # toplevel curses screen
sheet = None # current sheet

windowWidth = None
windowHeight = None

colors = {
    'bold': curses.A_BOLD,
    'reverse': curses.A_REVERSE,
    'normal': curses.A_NORMAL,
}

class attrdict(object):
    def __init__(self, d):
        self.__dict__ = d

class CommandPrefixException(Exception):
    pass

def error(s):
    raise Exception(s)

UnspecifiedType = lambda r=None: str(r)
UnspecifiedType.__name__ = ''

base_commands = collections.OrderedDict()
base_options = collections.OrderedDict()
options = attrdict(base_options)

def setup_sheet_commands():
    def command(ch, cmdstr, helpstr):
        base_commands[('', ch)] = (cmdstr, helpstr)
    def global_command(ch, cmdstr, helpstr):
        base_commands[('g', ch)] = (cmdstr, helpstr)
    def option(name, default, helpstr=''):
        base_options[name] = default
#        base_options[name] = (default, helpstr)
    def theme(name, default, helpstr=''):
        base_options[name] = default
#        base_options[name] = (default, helpstr)

    option('csv_dialect', 'excel', 'dialect passed to csv.reader')
    option('csv_delimiter', ',', 'delimiter passed to csv.reader')
    option('csv_quotechar', '"', 'quotechar passed to csv.reader')
    option('csv_header', '', 'parse first row of CSV as column names')

    option('debug', '', 'abort on error and display stacktrace')
    option('readonly', '', 'disable saving')

    option('encoding', 'utf-8', 'as passed to codecs.open')
    option('encoding_errors', 'surrogateescape', 'as passed to codecs.open')

    option('ColumnStats', False, 'include mean/median/etc on Column sheet')

    option('SubsheetSep', '~')

    # display/color scheme
    theme('SheetNameFmt', '%s| ', 'status line prefix')
    theme('ch_VisibleNone', '',  'visible contents of a cell whose value was None')
    theme('ch_FunctionError', '¿', 'when computation fails due to exception')
    theme('ch_Histogram', '*')
    theme('ch_ColumnFiller', ' ', 'pad chars after column value')
    theme('ch_LeftMore', '<')
    theme('ch_RightMore', '>')
    theme('ch_ColumnSep', '|', 'chars between columns')
    theme('ch_Ellipsis', '…')
    theme('ch_StatusSep', ' | ')
    theme('ch_KeySep', '/')
    theme('ch_EditPadChar', '_')
    theme('ch_Unprintable', '.')
    theme('ch_WrongType', '~')
    theme('ch_Error', '!')

    theme('c_default', 'normal')
    theme('c_Header', 'bold')
    theme('c_CurHdr', 'reverse')
    theme('c_CurRow', 'reverse')
    theme('c_CurCol', 'bold')
    theme('c_KeyCols', 'brown')
    theme('c_StatusLine', 'bold')
    theme('c_SelectedRow', 'green')
    theme('c_ColumnSep', 'blue')
    theme('c_EditCell', 'normal')
    theme('c_WrongType', 'magenta')
    theme('c_Error', 'red')

    command(Key.F1,    'vd.push(sheet.CommandHelp())', 'push help sheet')
    command(Key('q'),  'vd.sheets.pop(0)', 'pop this sheet')

    command(Key.LEFT,  'sheet.cursorRight(-1)', 'go one column left')
    command(Key.DOWN,  'sheet.cursorDown(+1)', 'go one row down')
    command(Key.UP,    'sheet.cursorDown(-1)', 'go one row up')
    command(Key.RIGHT, 'sheet.cursorRight(+1)', 'go one column right')
    command(Key.NPAGE, 'sheet.cursorDown(sheet.nVisibleRows); sheet.topRowIndex += sheet.nVisibleRows', 'scroll one page down')
    command(Key.PPAGE, 'sheet.cursorDown(-sheet.nVisibleRows); sheet.topRowIndex -= sheet.nVisibleRows', 'scroll one page up')
    command(Key.HOME,  'sheet.topRowIndex = sheet.cursorRowIndex = 0', 'go to top row')
    command(Key.END,   'sheet.cursorRowIndex = len(sheet.rows)-1', 'go to last row')

    command(Key('h'), 'sheet.cursorRight(-1)', 'go one column left')
    command(Key('j'), 'sheet.cursorDown(+1)', 'go one row down')
    command(Key('k'), 'sheet.cursorDown(-1)', 'go one row up')
    command(Key('l'), 'sheet.cursorRight(+1)', 'go one column right')

    command(Shift.H, 'sheet.moveVisibleCol(sheet.cursorVisibleColIndex, max(sheet.cursorVisibleColIndex-1, 0)); sheet.cursorVisibleColIndex -= 1', 'move this column one left')
    command(Shift.J, 'sheet.cursorRowIndex = moveListItem(sheet.rows, sheet.cursorRowIndex, min(sheet.cursorRowIndex+1, sheet.nRows-1))', 'move this row one down')
    command(Shift.K, 'sheet.cursorRowIndex = moveListItem(sheet.rows, sheet.cursorRowIndex, max(sheet.cursorRowIndex-1, 0))', 'move this row one up')
    command(Shift.L, 'sheet.moveVisibleCol(sheet.cursorVisibleColIndex, min(sheet.cursorVisibleColIndex+1, sheet.nVisibleCols-1)); sheet.cursorVisibleColIndex += 1', 'move this column one right')

    command(Ctrl.G, 'status(sheet.statusLine)', 'show this sheet info')
    command(Ctrl.P, 'status(vd.statusHistory[0])', 'show previous status line again')
    command(Ctrl.V, 'status(initialStatus)', 'show version information')

    command(Key('t'), 'sheet.topRowIndex = sheet.cursorRowIndex', 'scroll cursor row to top of screen')
    command(Key('m'), 'sheet.topRowIndex = sheet.cursorRowIndex-int(sheet.nVisibleRows/2)', 'scroll cursor row to middle of screen')
    command(Key('b'), 'sheet.topRowIndex = sheet.cursorRowIndex-sheet.nVisibleRows+1', 'scroll cursor row to bottom of screen')

    command(Key('<'), 'sheet.skipUp()', 'skip up this column to previous value')
    command(Key('>'), 'sheet.skipDown()', 'skip down this column to next value')

    command(Key('_'), 'sheet.cursorCol.width = sheet.cursorCol.getMaxWidth(sheet.visibleRows)', 'set this column width to fit visible cells')
    command(Key('-'), 'sheet.cursorCol.width = 0', 'hide this column')
    command(Key('^'), 'sheet.cursorCol.name = sheet.cursorCol.getDisplayValue(sheet.cursorRow)', 'set this column header to this cell value')
    command(Key('!'), 'sheet.toggleKeyColumn(sheet.cursorColIndex)', 'toggle this column as a key column')

    command(Key('@'), 'sheet.cursorCol.type = date', 'set column type to ISO8601 datetime')
    command(Key('#'), 'sheet.cursorCol.type = int', 'set column type to integer')
    command(Key('$'), 'sheet.cursorCol.type = str', 'set column type to string')
    command(Key('%'), 'sheet.cursorCol.type = float', 'set column type to float')
    command(Key('~'), 'sheet.cursorCol.type = detectType(sheet.cursorValue)', 'autodetect type of column by its data')

    command(Key('['), 'sheet.rows = sorted(sheet.rows, key=lambda r: sheet.cursorCol.getValue(r))', 'sort by this column ascending')
    command(Key(']'), 'sheet.rows = sorted(sheet.rows, key=lambda r: sheet.cursorCol.getValue(r), reverse=True)', 'sort by this column descending')
    command(Ctrl.E, 'options.debug = True; error(vd.lastErrors[-1])', 'abort and print last error to terminal')
    command(Ctrl.D, 'options.debug = not options.debug; status("debug " + ("ON" if options.debug else "OFF"))', 'toggle debug mode')

    command(Shift.E, 'if vd.lastErrors: vd.push(VSheetText("last_error", vd.lastErrors[-1]))', 'open stack trace for most recent error')
    command(Shift.F, 'vd.push(VSheetFreqTable(sheet, sheet.cursorCol))', 'open frequency table from values in this column')

    command(Key('d'), 'sheet.rows.pop(sheet.cursorRowIndex)', 'delete this row')

    command(Key('g'), 'raise CommandPrefixException("g")', 'add global prefix')

    command(Shift.S, 'vd.push(vd.sheets)', 'open Sheet stack')
    command(Shift.C, 'vd.push(sheet.columnsSheet)', 'open Columns for this sheet')
    command(Shift.O, 'vd.push(VSheetDict("options", options.__dict__))', 'open Options')

    command(Key('/'), 'sheet.searchRegex(inputLine(prompt="/"), columns=[sheet.cursorCol], moveCursor=True)', 'search this column forward for regex')
    command(Key('?'), 'sheet.searchRegex(inputLine(prompt="?"), columns=[sheet.cursorCol], backward=True, moveCursor=True)', 'search this column backward for regex')
    command(Key('n'), 'sheet.searchRegex(columns=[sheet.cursorCol], moveCursor=True)', 'go to next match')
    command(Key('p'), 'sheet.searchRegex(columns=[sheet.cursorCol], backward=True, moveCursor=True)', 'go to previous match')

    command(Key(' '), 'sheet.toggle([sheet.cursorRow]); sheet.cursorDown(1)', 'toggle select of this row')
    command(Key('s'), 'sheet.select([sheet.cursorRow]); sheet.cursorDown(1)', 'select this row')
    command(Key('u'), 'sheet.unselect([sheet.cursorRow]); sheet.cursorDown(1)', 'unselect this row')
    command(Key('|'), 'sheet.select(sheet.rows[r] for r in sheet.searchRegex(inputLine(prompt="|"), columns=[sheet.cursorCol]))', 'select rows by regex in this column')
    command(Key('\\'), 'sheet.unselect(sheet.rows[r] for r in sheet.searchRegex(inputLine(prompt="\\\\"), columns=[sheet.cursorCol]))', 'unselect rows by regex in this column')

    command(Shift.R, 'sheet.source.type = inputLine("change type to: ", value=initialsheet.source.type)', 'set source type of this sheet')
    command(Ctrl.R, 'open_source(vd.sheets.pop(0).source); status("reloaded")', 'reload sheet from source')
    command(Ctrl.S, 'saveSheet(sheet, inputLine("save to: "))', 'save this sheet to new file')
    command(Key('o'), 'open_source(inputLine("open: "))', 'open local file or url')
    command(Ctrl.O, 'expr = inputLine("eval: "); pushPyObjSheet(expr, eval(expr))', 'eval Python expression and open the result')

    command(Key('e'), 'sheet.cursorCol.setValue(sheet.cursorRow, sheet.editCell(sheet.cursorVisibleColIndex))', 'edit this cell')
    command(Key('c'), 'sheet.cursorVisibleColIndex = sheet.findColIdx(inputLine("goto column name: "), sheet.visibleCols)', 'goto visible column by name')
    command(Key('r'), 'sheet.cursorRowIndex = int(inputLine("goto row number: "))', 'goto row number')

    command(Key('='), 'sheet.addColumn(ColumnExpr(sheet, inputLine("new column expr=")), index=sheet.cursorColIndex+1)', 'add column by expr')
    command(Key(':'), 'sheet.addColumn(ColumnRegex(sheet, inputLine("new column regex:")), index=sheet.cursorColIndex+1)', 'add column by regex')
    command(Ctrl('^'), 'vd.sheets[0], vd.sheets[1] = vd.sheets[1], vd.sheets[0]', 'jump to previous sheet')
    command(Key.TAB,  'moveListItem(vd.sheets, 0, len(vd.sheets))', 'cycle through sheet stack')
    command(Key.BTAB, 'moveListItem(vd.sheets, -1, 0)', 'reverse cycle through sheet stack')

# when used with 'g' prefix
    global_command(Key('q'), 'vd.sheets.clear()', 'pop all sheets (clean exit)')

    global_command(Key('h'), 'sheet.cursorVisibleColIndex = sheet.leftVisibleColIndex = 0', 'go to leftmost column')
    global_command(Key('k'), 'sheet.cursorRowIndex = sheet.topRowIndex = 0', 'go to top row')
    global_command(Key('j'), 'sheet.cursorRowIndex = len(sheet.rows); sheet.topRowIndex = sheet.cursorRowIndex-sheet.nVisibleRows', 'go to bottom row')
    global_command(Key('l'), 'sheet.cursorVisibleColIndex = len(sheet.visibleCols)-1', 'go to rightmost column')

    global_command(Shift.H, 'moveListItem(sheet.columns, sheet.cursorColIndex, 0)', 'move this column all the way to the left')
    global_command(Shift.J, 'moveListItem(sheet.rows, sheet.cursorRowIndex, sheet.nRows)', 'move this row all the way to the bottom')
    global_command(Shift.K, 'moveListItem(sheet.rows, sheet.cursorRowIndex, 0)', 'move this row all the way to the top')
    global_command(Shift.L, 'moveListItem(sheet.columns, sheet.cursorColIndex, sheet.nCols)', 'move this column all the way to the right')

    global_command(Key('_'), 'for c in sheet.visibleCols: c.width = c.getMaxWidth(sheet.visibleRows)', 'set width of all columns to fit visible cells')
    global_command(Key('^'), 'for c in sheet.visibleCols: c.name = c.getDisplayValue(sheet.cursorRow)', 'set names of all visible columns to this row')
    global_command(Key('~'), 'for c in sheet.visibleCols: c.type = detectType(c.getValue(sheet.cursorRow))', 'autodetect types of all visible columns by their data')

    global_command(Shift.E, 'vd.push(VSheetText("last_error", "\\n\\n".join(vd.lastErrors)))', 'open last 10 errors')

    global_command(Key('/'), 'sheet.searchRegex(inputLine(prompt="/"), moveCursor=True, columns=sheet.visibleCols)', 'search regex forward in all visible columns')
    global_command(Key('?'), 'sheet.searchRegex(inputLine(prompt="?"), backward=True, moveCursor=True, columns=sheet.visibleCols)', 'search regex backward in all visible columns')
    global_command(Key('n'), 'sheet.cursorRowIndex = max(sheet.searchRegex())', 'go to first match')
    global_command(Key('p'), 'sheet.cursorRowIndex = min(sheet.searchRegex())', 'go to last match')

    global_command(Key(' '), 'sheet.toggle(sheet.rows)', 'toggle select of all rows')
    global_command(Key('s'), 'sheet.select(sheet.rows)', 'select all rows')
    global_command(Key('u'), 'sheet._selectedRows = {}', 'unselect all rows')

    global_command(Key('|'), 'sheet.select(sheet.rows[r] for r in sheet.searchRegex(inputLine(prompt="|"), columns=sheet.visibleCols))', 'select rows by regex in all visible columns')
    global_command(Key('\\'), 'sheet.unselect(sheet.rows[r] for r in sheet.searchRegex(inputLine(prompt="\\\\"), columns=sheet.visibleCols))', 'unselect rows by regex in all visible columns')

    global_command(Key('d'), 'sheet.rows = [r for r in sheet.rows if not sheet.isSelected(r)]; sheet._selectedRows = {}', 'delete all selected rows')

    global_command(Ctrl.P, 'vd.push(VSheetText("statuses", vd.statusHistory))', 'open last 100 statuses')

    # experimental commands
    command(Key('"'), 'sheet.rows.insert(sheet.cursorRowIndex, copy(sheet.cursorRow))', 'insert duplicate of this row')

# end setup_sheet_commands

### VisiData core

def moveListItem(L, fromidx, toidx):
    r = L.pop(fromidx)
    L.insert(toidx, r)
    return toidx

class date:
    def __init__(self, s):
        if isinstance(s, int) or isinstance(s, float):
            self.dt = datetime.datetime.fromtimestamp(s)
        elif isinstance(s, str):
            import dateutil.parser
            self.dt = dateutil.parser.parse(s)
        else:
            assert isinstance(s, datetime.datetime)
            self.dt = s

    def __str__(self):
        return self.dt.strftime('%Y-%m-%d %H:%M:%S')

    def __lt__(self, a):
        return self.dt < a.dt


def detectType(v):
    def tryType(T, v):
        try:
            v = T(v)
            return T
        except:
            return None

    return tryType(int, v) or tryType(float, v) or tryType(date, v) or str

class VSheet:
    def __init__(self, name, src=None):
        self.name = name
        self.source = src
        self.rows = []
        self.cursorRowIndex = 0  # absolute index of cursor into self.rows
        self.cursorVisibleColIndex = 0  # index of cursor into self.visibleCols

        self.topRowIndex = 0     # cursorRowIndex of topmost row
        self.leftVisibleColIndex = 0    # cursorVisibleColIndex of leftmost column

        # as computed during draw()
        self.rowLayout = {}      # [rowidx] -> y
        self.visibleColLayout = {}      # [vcolidx] -> (x, w)

        # all columns in display order
        self.columns = []
        self.nKeys = 0           # self.columns[:nKeys] are all pinned to the left and matched on join
        self._columnsSheet = None

        # current search term
        self.currentRegex = None
        self.currentRegexColumns = None

        self._selectedRows = {}   # id(row) -> row

        # specialized sheet keys
        self.commands = base_commands.copy()

    def __repr__(self):
        return self.name

    def isSelected(self, r):
        return id(r) in self._selectedRows

    def command(self, key, cmdstr, helpstr=''):
#        if key in self.commands:
#            status('overriding key %s' % key)
        self.commands[('', key)] = (cmdstr, helpstr)

    def findColIdx(self, colname, columns=None):
        if columns is None:
            columns = self.columns
        cols = list(colidx for colidx, c in enumerate(columns) if c.name == colname)
        if not cols:
            error('no column named "%s"' % colname)
        elif len(cols) > 1:
            status('%d columns named "%s"' % (len(cols), colname))
        return cols[0]

    @property
    def name(self):
        return self._name

    @name.setter
    def name(self, name):
        self._name = name.replace(' ', '_')

    @property
    def columnsSheet(self):
        if not self._columnsSheet:
            self._columnsSheet = VSheetColumns(self)
        return self._columnsSheet

    @property
    def nVisibleRows(self):
        return windowHeight-2

    @property
    def cursorCol(self):
        return self.visibleCols[self.cursorVisibleColIndex]

    @property
    def cursorRow(self):
        return self.rows[self.cursorRowIndex]

    @property
    def visibleRows(self):  # onscreen rows
        return self.rows[self.topRowIndex:self.topRowIndex+windowHeight-2]

    @property
    def visibleCols(self):  # non-hidden cols
        return [c for c in self.columns if not c.hidden]

    @property
    def cursorColIndex(self):
        return self.columns.index(self.cursorCol)

    @property
    def selectedRows(self):
        return [r for r in self.rows if id(r) in self._selectedRows]

    @property
    def keyCols(self):
        return self.columns[:self.nKeys]

    @property
    def keyColNames(self):
        return options.ch_KeySep.join(c.name for c in self.keyCols)

    @property
    def cursorValue(self):
        return self.cellValue(self.cursorRowIndex, self.cursorColIndex)

    @property
    def statusLine(self):
        return 'row %s/%s (%s selected); %d/%d columns visible' % (self.cursorRowIndex, len(self.rows), len(self._selectedRows), self.nVisibleCols, self.nCols)

    @property
    def nRows(self):
        return len(self.rows)

    @property
    def nCols(self):
        return len(self.columns)

    @property
    def nVisibleCols(self):
        return len(self.visibleCols)

    def moveVisibleCol(self, fromVisColIdx, toVisColIdx):
        fromColIdx = self.columns.index(self.visibleCols[fromVisColIdx])
        toColIdx = self.columns.index(self.visibleCols[toVisColIdx])
        moveListItem(self.columns, fromColIdx, toColIdx)
        return toVisColIdx

    def cursorDown(self, n):
        self.cursorRowIndex += n

    def cursorRight(self, n):
        self.cursorVisibleColIndex += n

    def cellValue(self, rownum, col):
        if not isinstance(col, VColumn):
            # assume it's the column number
            col = self.columns[col]
        return col.getValue(self.rows[rownum])

    def addColumn(self, col, index=None):
        if index is None:
            index = len(self.columns)
        if col:
            self.columns.insert(index, col)

    def toggleKeyColumn(self, colidx):
        if self.cursorColIndex >= self.nKeys: # if not a key, add it
            moveListItem(self.columns, self.cursorColIndex, self.nKeys)
            self.nKeys += 1
        else:  # otherwise move it after the last key
            self.nKeys -= 1
            moveListItem(self.columns, self.cursorColIndex, self.nKeys)

    def skipDown(self):
        pv = self.cursorValue
        for i in range(self.cursorRowIndex+1, len(self.rows)):
            if self.cellValue(i, self.cursorColIndex) != pv:
                self.cursorRowIndex = i
                return

        status('no different value down this column')

    def skipUp(self):
        pv = self.cursorValue
        for i in range(self.cursorRowIndex, -1, -1):
            if self.cellValue(i, self.cursorColIndex) != pv:
                self.cursorRowIndex = i
                return

        status('no different value up this column')

    def toggle(self, rows):
        for r in rows:
            if id(r) in self._selectedRows:
                del self._selectedRows[id(r)]
            else:
                self._selectedRows[id(r)] = r

    def select(self, rows):
        rows = list(rows)
        before = len(self._selectedRows)
        self._selectedRows.update(dict((id(r), r) for r in rows))
        status('selected %s/%s rows' % (len(self._selectedRows)-before, len(rows)))

    def unselect(self, rows):
        rows = list(rows)
        before = len(self._selectedRows)
        for r in rows:
            if id(r) in self._selectedRows:
                del self._selectedRows[id(r)]
        status('unselected %s/%s rows' % (before-len(self._selectedRows), len(rows)))

    def columnsMatch(self, row, columns, func):
        for c in columns:
            m = func(c.getDisplayValue(row))
            if m:
                return True
        return False

    def checkCursor(self):
        # keep cursor within actual available rowset
        if self.cursorRowIndex <= 0:
            self.cursorRowIndex = 0
        elif self.cursorRowIndex >= len(self.rows):
            self.cursorRowIndex = len(self.rows)-1

        if self.cursorVisibleColIndex <= 0:
            self.cursorVisibleColIndex = 0
        elif self.cursorVisibleColIndex >= self.nVisibleCols:
            self.cursorVisibleColIndex = self.nVisibleCols-1

        if self.topRowIndex <= 0:
            self.topRowIndex = 0
        elif self.topRowIndex > len(self.rows):
            self.topRowIndex = len(self.rows)-1

        # (x,y) is relative cell within screen viewport
        x = self.cursorVisibleColIndex - self.leftVisibleColIndex
        y = self.cursorRowIndex - self.topRowIndex + 1  # header

        # check bounds, scroll if necessary
        if y < 1:
            self.topRowIndex = self.cursorRowIndex
        elif y > windowHeight-2:
            self.topRowIndex = self.cursorRowIndex-windowHeight+3

        if x <= 0:
            self.leftVisibleColIndex = self.cursorVisibleColIndex
        else:
            while True:
                if self.leftVisibleColIndex == self.cursorVisibleColIndex: # not much more we can do
                    break
                self.calcColLayout()
                if self.cursorVisibleColIndex < min(self.visibleColLayout.keys()):
                    self.leftVisibleColIndex -= 1
                    continue
                elif self.cursorVisibleColIndex > max(self.visibleColLayout.keys()):
                    self.leftVisibleColIndex += 1
                    continue

                cur_x, cur_w = self.visibleColLayout[self.cursorVisibleColIndex]
                left_x, left_w = self.visibleColLayout[self.leftVisibleColIndex]
                if cur_x+cur_w < windowWidth: # current columns fit entirely on screen
                    break
                self.leftVisibleColIndex += 1

    def searchRegex(self, regex=None, columns=None, backward=False, moveCursor=False):
        'sets row index if moveCursor; otherwise returns list of row indexes'
        if regex:
            self.currentRegex = re.compile(regex, re.IGNORECASE)

        if not self.currentRegex:
            status('no regex')
            return []

        if columns:
            self.currentRegexColumns = columns

        if not self.currentRegexColumns:
            status('no columns given')
            return []

        if backward:
            rng = range(self.cursorRowIndex-1, -1, -1)
            rng2 = range(self.nRows-1, self.cursorRowIndex-1, -1)
        else:
            rng = range(self.cursorRowIndex+1, self.nRows)
            rng2 = range(0, self.cursorRowIndex+1)

        matchingRowIndexes = []

        for r in rng:
            if self.columnsMatch(self.rows[r], self.currentRegexColumns, self.currentRegex.search):
                if moveCursor:
                    self.cursorRowIndex = r
                    return r
                matchingRowIndexes.append(r)

        for r in rng2:
            if self.columnsMatch(self.rows[r], self.currentRegexColumns, self.currentRegex.search):
                if moveCursor:
                    self.cursorRowIndex = r
                    status('search wrapped')
                    return r
                matchingRowIndexes.append(r)

        status('%s matches for /%s/' % (len(matchingRowIndexes), self.currentRegex.pattern))

        return matchingRowIndexes

    def calcColLayout(self):
        self.visibleColLayout = {}
        x = 0
        for vcolidx in range(0, len(self.visibleCols)):
            col = self.visibleCols[vcolidx]
            if col.width is None:
                col.width = col.getMaxWidth(self.visibleRows)+len(options.ch_LeftMore)+len(options.ch_RightMore)
            if vcolidx < self.nKeys or vcolidx >= self.leftVisibleColIndex:  # visible columns
                self.visibleColLayout[vcolidx] = (x, min(col.width, windowWidth-x))
                x += col.width+len(options.ch_ColumnSep)
            if x > windowWidth-1:
                break

    def drawColHeader(self, vcolidx):
        # choose attribute to highlight column header
        if vcolidx == self.cursorVisibleColIndex:  # cursor is at this column
            hdrattr = colors[options.c_CurHdr]
        elif vcolidx < self.nKeys:
            hdrattr = colors[options.c_KeyCols]
        else:
            hdrattr = colors[options.c_Header]

        col = self.visibleCols[vcolidx]
        x, colwidth = self.visibleColLayout[vcolidx]

        # ANameTC
        typedict = {
                int: '#',
                str: '$',
                float: '%',
                date: '@',
                UnspecifiedType: ' ',
            }
        T = typedict.get(col.type, '?')
        N = ' ' + (col.name or defaultColNames[vcolidx])  # save room at front for LeftMore
        if len(N) > colwidth-1:
            N = N[:colwidth-len(options.ch_Ellipsis)] + options.ch_Ellipsis
        vd().clipdraw(0, x, N, hdrattr, colwidth)
        vd().clipdraw(0, x+colwidth-1, T, hdrattr, 1)

        if vcolidx == self.leftVisibleColIndex and vcolidx > self.nKeys:
            A = options.ch_LeftMore
            scr.addstr(0, x, A, colors[options.c_ColumnSep])

        C = options.ch_ColumnSep
        if x+colwidth+len(C) <= windowWidth:
            scr.addstr(0, x+colwidth, C, colors[options.c_ColumnSep])


    def draw(self):
        numHeaderRows = 1
        scr.erase()  # clear screen before every re-draw
        sepchars = options.ch_ColumnSep
        if not self.columns:
            return status('no columns')

        self.rowLayout = {}
        self.calcColLayout()
        for vcolidx, colinfo in sorted(self.visibleColLayout.items()):
            x, colwidth = colinfo
            if x < windowWidth:  # only draw inside window
                self.drawColHeader(vcolidx)

                y = numHeaderRows
                for rowidx in range(0, windowHeight-2):
                    if self.topRowIndex + rowidx >= len(self.rows):
                        break

                    self.rowLayout[self.topRowIndex+rowidx] = y

                    row = self.rows[self.topRowIndex + rowidx]

                    if self.topRowIndex + rowidx == self.cursorRowIndex:  # cursor at this row
                        attr = colors[options.c_CurRow]
                    elif vcolidx < self.nKeys:
                        attr = colors[options.c_KeyCols]
                    else:
                        attr = colors[options.c_default]

                    if self.isSelected(row):
                        attr |= colors[options.c_SelectedRow]

                    if vcolidx == self.cursorVisibleColIndex:  # cursor is at this column
                        attr |= colors[options.c_CurCol]

                    cellval = self.visibleCols[vcolidx].getDisplayValue(row, colwidth-1)
                    vd().clipdraw(y, x, options.ch_ColumnFiller + cellval, attr, colwidth)

                    if isinstance(cellval, CalcErrorStr):
                        vd().clipdraw(y, x+colwidth-1, options.ch_Error, colors[options.c_Error])
                    elif isinstance(cellval, WrongTypeStr):
                        vd().clipdraw(y, x+colwidth-1, options.ch_WrongType, colors[options.c_WrongType])

                    if x+colwidth+len(sepchars) <= windowWidth:
                       scr.addstr(y, x+colwidth, sepchars, attr or colors[options.c_ColumnSep])

                    y += 1

        if vcolidx+1 < self.nVisibleCols:
            scr.addstr(0, windowWidth-1, options.ch_RightMore, colors[options.c_ColumnSep])

    def editCell(self, vcolidx=None):
        if vcolidx is None:
            vcolidx = self.cursorVisibleColIndex
        x, w = self.visibleColLayout[vcolidx]
        y = self.rowLayout[self.cursorRowIndex]

        currentValue = self.cellValue(self.cursorRowIndex, vcolidx)
        r = vd().editText(y, x, w, value=currentValue, fillchar=options.ch_EditPadChar)
        return self.visibleCols[vcolidx].type(r)  # convert input to column type

    def CommandHelp(self):
        vs = VSheet(self.name + '_help', self)
        vs.rows = list(self.commands.items())
        vs.columns = [VColumn('key', str, lambda r: r[0][0] + keyname(r[0][1])),
                VColumn('action', str, lambda r: r[1][1]),
                VColumn('global_action', str, lambda r,sheet=self: sheet.commands.get(('g', r[0][1]), ('', '-'))[1])]
        return vs

# end VSheet class

class VisiData:
    def __init__(self):
        self.sheets = VSheetSheets()
        self.statusHistory = []
        self._status = []
        self.status(initialStatus)
        self.lastErrors = []

    def status(self, s):
        strs = str(s)
        self._status.append(strs)
        self.statusHistory.insert(0, strs)
        del self.statusHistory[100:]  # keep most recent 100 only
        return s

    def displayRightStatus(self, prefixes, ch):
        rstat = "%s" % (keyname(ch)) # (chr(ch) if chr(ch).isprintable() else keyname(ch)
        self.clipdraw(windowHeight-1, windowWidth-len(rstat)-2, rstat, colors[options.c_StatusLine])

    def editText(self, y, x, w, **kwargs):
        v = editText(scr, y, x, w, **kwargs)
        self.status('"%s"' % v)
        return v

    def exceptionCaught(self, status=True):
        import traceback
        self.lastErrors.append(traceback.format_exc().strip())
        self.lastErrors = self.lastErrors[-10:]  # keep most recent
        if status:
            self.status(self.lastErrors[-1].splitlines()[-1])
        if options.debug:
            raise

    def run(self):
        global sheet
        global windowHeight, windowWidth
        windowHeight, windowWidth = scr.getmaxyx()
        ch = 32

        command_overrides = None
        prefixes = ''
        while True:
            if not self.sheets:
                # if no more sheets, exit
                return

            sheet = self.sheets[0]
            if sheet.nRows == 0:
                self.status('no rows')

            try:
                sheet.draw()
            except Exception as e:
                self.exceptionCaught()

            # draw status on last line
            attr = colors[options.c_StatusLine]
            statusstr = options.SheetNameFmt % sheet.name + options.ch_StatusSep.join(self._status)
            self.clipdraw(windowHeight-1, 0, statusstr, attr, windowWidth)
            self._status = []

            self.displayRightStatus(prefixes, ch)  # visible for this getch

            curses.doupdate()
            ch = scr.getch()

            self.displayRightStatus(prefixes, ch)  # visible for commands that wait with getch

            if ch == curses.KEY_RESIZE:
                windowHeight, windowWidth = scr.getmaxyx()
            elif ch == curses.KEY_MOUSE:
                try:
                    devid, x, y, z, bstate = curses.getmouse()
                    sheet.cursorRowIndex = sheet.topRowIndex+y-1
                except Exception:
                    self.exceptionCaught()
            elif (prefixes, ch) in sheet.commands:
                cmdstr, helpstr = sheet.commands.get((prefixes, ch))
                try:
                    exec(cmdstr, globals(), { 'vd': vd() })
                    prefixes = ''
                except CommandPrefixException as e:
                    prefixes += str(e)
                except EscapeException as e:  # user aborted
                    prefixes = ''
                    self.status(keyname(e.args[0]))
                except Exception:
                    prefixes = ''
                    self.exceptionCaught()
                    self.status(cmdstr)
            else:
                prefixes = ''
                self.status('no command for key "%s" (%d) with prefixes "%s"' % (keyname(ch), ch, prefixes))

            sheet.checkCursor()

    def push(self, vs):
        if vs:
            if vs in self.sheets:
                self.sheets.remove(vs)
            self.sheets.insert(0, vs)
            return vs

    def clipdraw(self, y, x, s, attr=curses.A_NORMAL, w=None):
        s = s.replace('\n', '\\n')
        try:
            if w is None:
                w = windowWidth-1
            w = min(w, windowWidth-x-1)
            if w == 0:  # no room anyway
                return

            # convert to string just before drawing
            s = str(s)
            if len(s) > w:
                scr.addstr(y, x, s[:w-1] + options.ch_Ellipsis, attr)
            else:
                scr.addstr(y, x, s, attr)
                if len(s) < w:
                    scr.addstr(y, x+len(s), options.ch_ColumnFiller*(w-len(s)), attr)
        except Exception as e:
            self.status('clipdraw error: y=%s x=%s len(s)=%s w=%s' % (y, x, len(s), w))
            self.exceptionCaught()
# end VisiData class

class WrongTypeStr(str):
    pass

class CalcErrorStr(str):
    pass

## columns
class VColumn:
    def __init__(self, name, type=UnspecifiedType, func=lambda r: r, width=None):
        self.name = name
        self.func = func
        self.width = width
        self.type = type
        self.expr = None  # Python string expression if computed column
        self._fmtstr = None

    @property
    def name(self):
        return self._name

    @name.setter
    def name(self, name):
        self._name = name.replace(' ', '_')

    @property
    def fmtstr(self):
        if self._fmtstr is not None: return self._fmtstr
        elif self.type is int: return '%d'
        elif self.type is float: return '%.02f'
        else: return '%s'

    @property
    def hidden(self):
        return self.width == 0

    def nEmpty(self, rows):
        vals = self.values(rows)
        return sum(1 for v in vals if v == '' or v == None)

    def values(self, rows):
        return [self.getValue(r) for r in rows]

    def getValue(self, row):
        try:
            v = self.func(row)
        except Exception as e:
            vd().exceptionCaught(status=False)
            raise

        try:
            return self.type(v)  # convert type on-the-fly
        except Exception as e:
            vd().exceptionCaught(status=False)
            return self.type()  # return a suitable value for this type

    def getDisplayValue(self, row, width=None):
        try:
            cellval = self.func(row)
        except Exception as e:
            vd().exceptionCaught(status=False)
            return CalcErrorStr(options.ch_FunctionError)

        if cellval is None:
            return options.ch_VisibleNone

        if isinstance(cellval, bytes):
            cellval = cellval.decode(options.encoding)

        try:
            cellval = self.fmtstr % self.type(cellval)  # convert type on-the-fly
            if width and self.type in (int, float): cellval = cellval.rjust(width-1)
        except Exception as e:
            vd().exceptionCaught(status=False)
            cellval = WrongTypeStr(str(cellval))

        return cellval

    def setValue(self, row, value):
        if hasattr(self.func, 'setter'):
            self.func.setter(row, value)
        else:
            error('column cannot be changed')

    def getMaxWidth(self, rows):
        if len(rows) == 0:
            return 0

        return min(max(max(len(self.getDisplayValue(r)) for r in rows), len(self.name)), int(windowWidth-3))+2


### common column setups and helpers
def setter(r, k, v):  # needed for use in lambda
    r[k] = v

def lambda_colname(colname):
    func = lambda r: r[colname]
    func.setter = lambda r,v,b=colname: setter(r,b,v)
    return func

def lambda_col(b):
    func = lambda r,b=b: r[b]
    func.setter = lambda r,v,b=b: setter(r,b,v)
    return func

def lambda_getattr(b):
    func = lambda r: getattr(r,b)
    func.setter = lambda r,v,b=b: setattr(r,b,v)
    return func

def lambda_subrow_wrap(func, subrowidx):
    'wraps original func to be func(r[subrowidx])'
    subrow_func = lambda r,i=subrowidx,f=func: r[i] and f(r[i]) or None
    subrow_func.setter = lambda r,v,i=subrowidx,f=func: r[i] and f.setter(r[i], v) or None
    return subrow_func

class OnDemandDict:
    def __init__(self, sheet, row):
        self.row = row
        self.sheet = sheet
    def keys(self):
        return [c.name for c in self.sheet.columns]
    def __getitem__(self, colname):
        colidx = self.sheet.findColIdx(colname, self.sheet.columns)
        return self.sheet.columns[colidx].getValue(self.row)


def evalCol(sheet, col, row):
    return eval(col.expr, {}, OnDemandDict(sheet, row))

def ColumnExpr(sheet, expr):
    if expr:
        vc = VColumn(expr)
        vc.expr = expr
        vc.func = lambda r,newcol=vc,sheet=sheet: evalCol(sheet, newcol, r)
        return vc

def getTransformedValue(oldval, searchregex, replregex):
    m = re.search(searchregex, oldval)
    if not m:
        return None
    return m.expand(replregex)

def ColumnRegex(sheet, regex):
    if regex:
        vc = VColumn(regex)
        vc.expr, vc.searchregex, vc.replregex = regex.split('/')   # TODO: better syntax
        vc.func = lambda r,newcol=vc,sheet=sheet: getTransformedValue(str(evalCol(sheet, newcol, r)), newcol.searchregex, newcol.replregex)
        return vc

def getPublicAttrs(obj):
    return [k for k in dir(obj) if not k.startswith('_') and not callable(getattr(obj, k))]

def PyobjColumns(exampleRow):
    'columns for each public attribute on an object'
    return [VColumn(k, type(getattr(exampleRow, k)), lambda_getattr(k)) for k in getPublicAttrs(exampleRow)]

def ArrayColumns(n):
    'columns that display r[0]..r[n]'
    return [VColumn('', UnspecifiedType, lambda_col(colnum)) for colnum in range(n)]

def ArrayNamedColumns(columns):
    'columns is a list of column names, mapping to r[0]..r[n]'
    return [VColumn(colname, UnspecifiedType, lambda_col(i)) for i, colname in enumerate(columns)]

def AttrColumns(colnames):
    'colnames is list of attribute names'
    return [VColumn(name, UnspecifiedType, lambda_getattr(name)) for name in colnames]

### sheet layouts
#### generic list/dict/object browsing
def pushPyObjSheet(name, pyobj, src=None):
    if isinstance(pyobj, list):
        return vd().push(VSheetList(name, pyobj, src=src))
    elif isinstance(pyobj, dict):
        return vd().push(VSheetDict(name, pyobj))
    elif isinstance(pyobj, object):
        return vd().push(VSheetObject(name, pyobj))
    else:
        status('unknown type ' + type(pyobj))

class VSheetList(VSheet):
    def __init__(self, name, obj, columns=None, src=None):
        'columns is a list of strings naming attributes on the objects within the obj'
        super().__init__(name, src or obj)
        assert isinstance(obj, list)
        self.rows = obj
        if columns:
            self.columns = AttrColumns(columns)
        elif isinstance(obj[0], dict):  # list of dict
            self.columns = [VColumn(k, type(obj[0][k]), lambda_colname(k)) for k in obj[0].keys()]
            self.nKeys = 1
        else:
            self.columns = [VColumn(name)]
        self.command(Key.ENTER, 'pushPyObjSheet("%s[%s]" % (sheet.name, sheet.cursorRowIndex), sheet.cursorRow).cursorRowIndex = sheet.cursorColIndex', 'dive into this row')

class VSheetDict(VSheet):
    def __init__(self, name, mapping):
        super().__init__(name, mapping)
        self.rows = list(list(x) for x in mapping.items())
        self.columns = [
            VColumn('key', str, lambda_col(0)),
            VColumn('value', UnspecifiedType, lambda_col(1)) ]
        self.nKeys = 1
        self.command(Key.ENTER, 'if sheet.cursorColIndex == 1: pushPyObjSheet(sheet.name + options.SubsheetSep + sheet.cursorRow[0], sheet.cursorRow[1])', 'dive into this value')
        self.command(Key('e'), 'sheet.source[sheet.cursorRow[0]] = sheet.cursorRow[1] = sheet.editCell(1)', 'edit this value')

class VSheetObject(VSheet):
    def __init__(self, name, obj):
        super().__init__(name, obj)
        self.command(Key.ENTER, 'pushPyObjSheet(sheet.name + options.SubsheetSep + sheet.cursorRow[0], sheet.cursorRow[1]() if callable(sheet.cursorRow[1]) else sheet.cursorRow[1])', 'dive into this value')
        self.command(Key('e'), 'setattr(sheet.source, sheet.cursorRow[0], sheet.editCell(1)); sheet.reload()', 'edit this value')
        self.reload()

    def reload(self):
        valfunc = lambda_col(1)
        valfunc.setter = lambda r,v,obj=self.source: setattr(obj, r[0], v)
        self.columns = [
            VColumn(type(self.source).__name__ + '_attr', str, lambda_col(0)),
            VColumn('value', UnspecifiedType, valfunc) ]
        self.nKeys = 1
        self.rows = [(k, getattr(self.source, k)) for k in dir(self.source)]

#### specialized meta sheets
class VSheetSheets(VSheet, list):
    def __init__(self):
        VSheet.__init__(self, 'sheets', 'vd.sheets')
        self.command(Key.ENTER,    'moveListItem(sheet, sheet.cursorRowIndex, 0); vd.sheets.pop(1)', 'go to this sheet')
        self.command(Key('&'), 'vd.sheets[0] = VSheetJoin(sheet.selectedRows, jointype="&")', 'open inner join of selected sheets')
        self.command(Key('+'), 'vd.sheets[0] = VSheetJoin(sheet.selectedRows, jointype="+")', 'open outer join of selected sheets')
        self.command(Key('*'), 'vd.sheets[0] = VSheetJoin(sheet.selectedRows, jointype="*")', 'open full join of selected sheets')
        self.command(Key('~'), 'vd.sheets[0] = VSheetJoin(sheet.selectedRows, jointype="~")', 'open diff join of selected sheets')
        self.reload()

    def reload(self):
        self.columns = AttrColumns('name nRows nCols cursorValue keyColNames source'.split())
        self.nKeys = 1
        self.rows = self

class VSheetColumns(VSheet):
    def __init__(self, srcsheet):
        super().__init__(srcsheet.name + '_columns', srcsheet)
# on the Columns sheet, these affect the 'row' (column in the source sheet)
        self.command(Key('@'), 'sheet.cursorRow.type = date; sheet.cursorDown(+1)', 'set column type to datetime')
        self.command(Key('#'), 'sheet.cursorRow.type = int; sheet.cursorDown(+1)', 'set source column type to integer')
        self.command(Key('$'), 'sheet.cursorRow.type = str; sheet.cursorDown(+1)', 'set source column type to string')
        self.command(Key('%'), 'sheet.cursorRow.type = float; sheet.cursorDown(+1)', 'set source column type to decimal numeric type')
        self.command(Key('~'), 'sheet.cursorRow.type = detectType(sheet.cursorRow.getValue(sheet.source.cursorRow)); sheet.cursorDown(+1)', 'autodetect type of source column using its data')
        self.command(Key('!'), 'sheet.source.toggleKeyColumn(sheet.cursorRowIndex)', 'toggle key column on source sheet')
        self.command(Key('-'), 'sheet.cursorRow.width = 0', 'hide column on source sheet')
        self.command(Key('_'), 'sheet.cursorRow.width = sheet.cursorRow.getMaxWidth(sheet.source.rows)', 'set source column width to max width of its rows')
        self.reload()

    def reload(self):
        self.rows = self.source.columns
        self.nKeys = 1
        self.columns = [
            VColumn('column',   str, lambda_getattr('name')),
            VColumn('width',  int, lambda_getattr('width')),
            VColumn('type',   str, lambda r: r.type.__name__),
            VColumn('fmtstr', str, lambda_getattr('_fmtstr')),
            VColumn('expr',   str, lambda_getattr('expr')),
            VColumn('value',  UnspecifiedType, lambda c,sheet=sheet: c.getValue(sheet.cursorRow)),
#            VColumn('nulls',  int, lambda c,sheet=sheet: c.nEmpty(sheet.rows)),

#            VColumn('uniques',  int, lambda c,sheet=sheet: len(set(c.values(sheet.rows))), width=0),
#            VColumn('mode',   UnspecifiedType, lambda c: statistics.mode(c.values(sheet.rows)), width=0),
#            VColumn('min',    UnspecifiedType, lambda c: min(c.values(sheet.rows)), width=0),
#            VColumn('median', UnspecifiedType, lambda c: statistics.median(c.values(sheet.rows)), width=0),
#            VColumn('mean',   float, lambda c: statistics.mean(c.values(sheet.rows)), width=0),
#            VColumn('max',    UnspecifiedType, lambda c: max(c.values(sheet.rows)), width=0),
#            VColumn('stddev', float, lambda c: statistics.stdev(c.values(sheet.rows)), width=0),
            ]

#### slicing and dicing
class VSheetJoin(VSheet):
    def __init__(self, sheets, jointype='&'):
        super().__init__(jointype.join(vs.name for vs in sheets))
        self.source = sheets
        self.jointype = jointype
        self.reload()

    def reload(self):
        sheets = self.source
        # first element in joined row is the tuple of keys
        self.columns = []
        for colnum in range(sheets[0].nKeys):
            c = sheets[0].columns[colnum]
            self.columns.append(VColumn(c.name, c.type, lambda_subrow_wrap(lambda_col(colnum), 0)))
        self.nKeys = sheets[0].nKeys

        rowsBySheetKey = {}
        rowsByKey = {}

        for vs in sheets:
            rowsBySheetKey[vs] = {}
            for r in vs.rows:
                key = tuple(c.getValue(r) for c in vs.keyCols)
                rowsBySheetKey[vs][key] = r

        for sheetnum, vs in enumerate(sheets):
            # subsequent elements are the rows from each source, in order of the source sheets
            self.columns.extend(VColumn(c.name, c.type, lambda_subrow_wrap(c.func, sheetnum+1), c.width) for c in vs.columns[vs.nKeys:])
            for r in vs.rows:
                key = tuple(c.getValue(r) for c in vs.keyCols)
                if key not in rowsByKey:
                    rowsByKey[key] = [key] + [rowsBySheetKey[vs2].get(key) for vs2 in sheets]  # combinedRow

        if self.jointype == '&':  # inner join  (only rows with matching key on all sheets)
            self.rows = list(combinedRow for k, combinedRow in rowsByKey.items() if all(combinedRow))
        elif self.jointype == '+':  # outer join (all rows from first sheet)
            self.rows = list(combinedRow for k, combinedRow in rowsByKey.items() if combinedRow[1])
        elif self.jointype == '*':  # full join (keep all rows from all sheets)
            self.rows = list(combinedRow for k, combinedRow in rowsByKey.items())
        elif self.jointype == '~':  # diff join (only rows without matching key on all sheets)
            self.rows = list(combinedRow for k, combinedRow in rowsByKey.items() if not all(combinedRow))


class VSheetFreqTable(VSheet):
    def __init__(self, sheet, col):
        fqcolname = '%s_%s_freq' % (sheet.name, col.name)
        super().__init__(fqcolname, sheet)

        self.origCol = col
        self.values = collections.defaultdict(list)
        for r in sheet.rows:
            self.values[str(col.getValue(r))].append(r)
        self.reload()

    def reload(self):
        self.rows = sorted(self.values.items(), key=lambda r: len(r[1]), reverse=True)  # sort by num reverse
        self.largest = len(self.rows[0][1])+1

        self.columns = [
            VColumn(self.origCol.name, self.origCol.type, lambda_col(0)),
            VColumn('num', int, lambda r: len(r[1])),
            VColumn('percent', float, lambda r: len(r[1])*100/sheet.source.nRows),
            VColumn('histogram', str, lambda r,s=self: options.ch_Histogram*int(len(r[1])*80/s.largest), width=80)
        ]
        self.nKeys = 1

        self.command(Key(' '), 'sheet.source.toggle(sheet.cursorRow[1])', 'toggle these entries')
        self.command(Key('s'), 'sheet.source.select(sheet.cursorRow[1])', 'select these entries')
        self.command(Key('u'), 'sheet.source.unselect(sheet.cursorRow[1])', 'unselect these entries')
        self.command(Key.ENTER, 'vd.push(copy(sheet.source)).rows = sheet.values[str(sheet.columns[0].getValue(sheet.cursorRow))]', 'push new sheet with only this value')


### input formats and helpers

sourceCache = {}

class Path:
    def __init__(self, fqpn):
        self.fqpn = fqpn
        self.name = os.path.split(fqpn)[-1]
        self.suffix = os.path.splitext(self.name)[1][1:]
    def read_text(self):
        return open(self.resolve(), encoding=options.encoding, errors=options.encoding_errors).read()
    def is_dir(self):
        return os.path.isdir(self.resolve())
    def iterdir(self):
        return [self.parent] + [Path(os.path.join(self.fqpn, f)) for f in os.listdir(self.resolve())]
    def stat(self):
        return os.stat(self.resolve())
    def resolve(self):
        return os.path.expandvars(os.path.expanduser(self.fqpn))
    @property
    def parent(self):
        return Path(self.fqpn + "/..")
    def __str__(self):
        return self.fqpn

def getTextContents(p):
    if not p in sourceCache:
        sourceCache[p] = p.read_text()
    return sourceCache[p]

def open_source(p):
    if isinstance(p, Path):
        if p.is_dir():
            vs = VSheetDirectory(p)
        else:
            vs = globals().get('open_' + p.suffix, open_txt)(p)
    elif '://' in p:
        vs = openUrl(p)
    else:
        return open_source(Path(p))
    if vs:
        status('opened %s' % p.name)
    if isinstance(vs, VSheet):
        return vd().push(vs)

class VSheetText(VSheet):
    def __init__(self, name, content, src=None):
        super().__init__(name, src)
        self.columns = [VColumn(name, str)]
        if isinstance(content, list):
            self.rows = content
        elif isinstance(content, str):
            self.rows = content.split('\n')
        else:
            error('unknown text type ' + str(type(content)))

class VSheetDirectory(VSheet):
    def __init__(self, p):
        super().__init__(p.name, p)
        self.columns = [VColumn('filename', str, lambda r: r[0].name),
                        VColumn('type', str, lambda r: r[0].is_dir() and '/' or r[0].suffix),
                        VColumn('size', int, lambda r: r[1].st_size),
                        VColumn('mtime', date, lambda r: r[1].st_mtime)]
        self.command(Key.ENTER, 'open_source(sheet.cursorRow[0])', 'open file')  # path, filename
        self.reload()

    def reload(self):
        self.rows = [(p, p.stat()) for p in self.source.iterdir()]  #  if not p.name.startswith('.')]

def open_txt(p):
    contents = getTextContents(p)
    if '\t' in contents[:32]:
        return open_tsv(p)  # TSV often have .txt extension
    return VSheetText(p.name, contents, p)

class open_csv(VSheet):
    def __init__(self, p):
        super().__init__(p.name, p)
        contents = getTextContents(p)

        if options.csv_dialect == 'sniff':
            headers = contents[:1024]
            dialect = csv.Sniffer().sniff(headers)
            status('sniffed csv_dialect as %s' % dialect)
        else:
            dialect = options.csv_dialect

        rdr = csv.reader(io.StringIO(contents, newline=''), dialect=dialect, delimiter=options.csv_delimiter, quotechar=options.csv_quotechar)
        self.rows = [r for r in rdr]
        if options.csv_header:
            self.columns = ArrayNamedColumns(self.rows[0])
            self.rows = self.rows[1:]
        else:
            self.columns = ArrayColumns(len(self.rows[0]))

class open_tsv(VSheet):
    def __init__(self, p):
        super().__init__(p.name, p)
        lines = getTextContents(p).splitlines()

        if options.csv_header:
            self.columns = ArrayNamedColumns(lines[0].split('\t'))
            lines = lines[1:]
        else:
            self.columns = ArrayColumns(len(lines[0].split('\t')))

        self.rows = [L.split('\t') for L in lines]  # [rownum] -> [ field, ... ]

def open_json(p):
    import json
    pushPyObjSheet(p.name, json.loads(getTextContents(p)))

#### .xlsx
class open_xlsx(VSheet):
    def __init__(self, path):
        super().__init__(path.name, path)
        import openpyxl
        self.workbook = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        self.rows = list(self.workbook.sheetnames)
        self.columns = [VColumn('name', str)]
        self.command(Key.ENTER, 'vd.push(sheet.getSheet(sheet.cursorRow))', 'open this sheet')

    def getSheet(self, sheetname):
        'create actual VSheet from xlsx sheet'
        worksheet = self.workbook.get_sheet_by_name(sheetname)
        vs = VSheet('%s%s%s' % (self.source, options.SubsheetSep, sheetname), worksheet)
        vs.columns = ArrayColumns(worksheet.max_column)
        vs.rows = [ [cell.value for cell in row] for row in worksheet.iter_rows()]
        return vs

#### .hdf5
class VSheetH5Obj(VSheet):
    def __init__(self, name, hobj, src):
        super().__init__(name, src)
        self.hobj = hobj
        self.reload()

    def reload(self):
        import h5py
        if isinstance(self.hobj, h5py.Group):
            self.rows = [ self.hobj[objname] for objname in self.hobj.keys() ]
            self.columns = [
                VColumn(self.hobj.name, str, lambda r: r.name.split('/')[-1]),
                VColumn('type', str, lambda r: type(r).__name__),
                VColumn('nItems', int, lambda r: len(r)),
            ]
            self.command(Key.ENTER, 'vd.push(VSheetH5Obj(sheet.name+options.SubsheetSep+sheet.cursorRow.name, sheet.cursorRow, sheet.source))', 'open this group or dataset')
            self.command(Key('A'), 'vd.push(VSheetDict(sheet.cursorRow.name + "_attrs", sheet.cursorRow.attrs))', 'open metadata sheet for this object')
        elif isinstance(self.hobj, h5py.Dataset):
            if len(self.hobj.shape) == 1:
                self.rows = self.hobj[:]  # copy
                self.columns = [VColumn(colname, str, lambda_colname(colname)) for colname in self.hobj.dtype.names]
            elif len(self.hobj.shape) == 2:  # matrix
                self.rows = self.hobj[:]  # copy
                self.columns = ArrayColumns(self.hobj.shape[1])
            else:
                status('too many dimensions in shape %s' % str(self.hobj.shape))
        else:
            status('unknown h5 object type %s' % type(self.hobj))

class open_hdf5(VSheetH5Obj):
    def __init__(self, p):
        import h5py
        super().__init__(p.name, h5py.File(str(p), 'r'), p)
open_h5 = open_hdf5


class open_zip(VSheet):
    def __init__(self, p):
        import zipfile
        super().__init__(p.name, p)
        self.zfp = zipfile.ZipFile(p.fqpn, 'r')
        self.rows = self.zfp.infolist()
        self.columns = AttrColumns("filename file_size date_time compress_size".split())
        self.command(Ctrl.J, 'vd.push(sheet.open(sheet.cursorRow))', 'open this file')

    def open(self, zi):
        cachefn = zi.filename
        if not os.path.exists(cachefn):
            self.zfp.extract(zi)
        return open_source(cachefn)

#### databases
class VSheetBlaze(VSheet):
    def __init__(self, name, data, src):
        super().__init__(name, src)
        self.columns = ArrayNamedColumns(data.fields)
        self.rows = list(data)

def openUrl(url):
    m = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
    if m:
        return open_gspreadsheet(Path(m.group(1)))

    import blaze
    import datashape; datashape.coretypes._canonical_string_encodings.update({"utf8_unicode_ci": "U8"})
    fp = blaze.data(url)
    vs = VSheetList(url, [getattr(fp, tblname) for tblname in fp.fields], url)
    vs.command(Key.ENTER, 'vd.push(VSheetBlaze(sheet.cursorRow.name, sheet.cursorRow, sheet))', 'open this table')
    return vs

#### Google Sheets; requires credentials to be setup already

@functools.lru_cache()
def google_sheets():
    import httplib2
    import os

    from apiclient import discovery
    from oauth2client.file import Storage

    home_dir = os.path.expanduser('~')
    credential_dir = os.path.join(home_dir, '.credentials')
    credential_path = os.path.join(credential_dir, 'sheets.googleapis.com-python-quickstart.json')

    store = Storage(credential_path)
    credentials = store.get()
    if not credentials:
        status('Credentials required')
    elif credentials.invalid:
        status('No or Invalid credentials')
    else:
        http = credentials.authorize(httplib2.Http())
        discoveryUrl = ('https://sheets.googleapis.com/$discovery/rest?version=v4')
        service = discovery.build('sheets', 'v4', http=http, discoveryServiceUrl=discoveryUrl)

        return service.spreadsheets()

def open_gspreadsheet(p):
    sheets = google_sheets()
    sheet_md = sheets.get(spreadsheetId=p.name).execute()
    vs = VSheet(sheet_md['properties']['title'], p)
    vs.columns = [VColumn('title', lambda_eval('properties.title')),
                  VColumn('rowCount', lambda_eval('properties.gridProperties.rowCount')),
                  VColumn('columnCount', lambda_eval('properties.gridProperties.columnCount'))]
    vs.rows = sheet_md['sheets']
    vs.command(Key.ENTER, 'sheet.cursorRow')
    return vs

def open_gsheet(p):
    sheets = google_sheets()
    sheet = sheets.values().get(spreadsheetId=p.name).execute()
    pushPyObjSheet(p.name, sheet, p)
#    vs = VSheet(sheet_md['properties']['title'], p)
#    vs.columns = [VColumn('title', lambda_eval('properties.title')),
#                  VColumn('rowCount', lambda_eval('properties.gridProperties.rowCount')),
#                  VColumn('columnCount', lambda_eval('properties.gridProperties.columnCount'))]
#    vs.rows = sheet_md['sheets']
#    vs.command(Key.ENTER, 'sheet.cursorRow')
#    return vs

#### external addons
def open_py(p):
    contents = getTextContents(p)
    exec(contents, globals())
    status('executed %s' % p)

### Sheet savers

def saveSheet(sheet, fn):
    if options.readonly:
        status('readonly mode')
        return
    basename, ext = os.path.splitext(fn)
    funcname = 'save_' + ext[1:]
    globals().get(funcname, save_tsv)(sheet, fn)
    status('saved to ' + fn)

def save_tsv(sheet, fn):
    with open(fn, 'w', encoding=options.encoding, errors=options.encoding_errors) as fp:
        colhdr = '\t'.join(col.name for col in sheet.visibleCols) + '\n'
        if colhdr.strip():  # is anything but whitespace
            fp.write(colhdr)
        for r in sheet.rows:
            fp.write('\t'.join(col.getDisplayValue(r) for col in sheet.visibleCols) + '\n')

def save_csv(sheet, fn):
    with open(fn, 'w', newline='', encoding=options.encoding, errors=options.encoding_errors) as fp:
        cw = csv.writer(fp, dialect=options.csv_dialect, delimiter=options.csv_delimiter, quotechar=options.csv_quotechar)
        colnames = [col.name for col in sheet.visibleCols]
        if ''.join(colnames):
            cw.writerow(colnames)
        for r in sheet.rows:
            cw.writerow([col.getDisplayValue(r) for col in sheet.visibleCols])

### curses, options, init

def inputLine(prompt, value=''):
    'add a prompt to the bottom of the screen and get a line of input from the user'
    scr.addstr(windowHeight-1, 0, prompt)
    return vd().editText(windowHeight-1, len(prompt), windowWidth-len(prompt)-8, value=value, attr=colors[options.c_EditCell], unprintablechar=options.ch_Unprintable)


nextColorPair = 1
def setupcolors(stdscr, f, *args):
    def makeColor(fg, bg):
        global nextColorPair
        if curses.has_colors():
            curses.init_pair(nextColorPair, fg, bg)
            c = curses.color_pair(nextColorPair)
            nextColorPair += 1
        else:
            c = curses.A_NORMAL

        return c

    curses.meta(1)  # allow "8-bit chars"

    colors['red'] = curses.A_BOLD | makeColor(curses.COLOR_RED, curses.COLOR_BLACK)
    colors['blue'] = curses.A_BOLD | makeColor(curses.COLOR_BLUE, curses.COLOR_BLACK)
    colors['green'] = curses.A_BOLD | makeColor(curses.COLOR_GREEN, curses.COLOR_BLACK)
    colors['brown'] = makeColor(curses.COLOR_YELLOW, curses.COLOR_BLACK)
    colors['yellow'] = curses.A_BOLD | colors['brown']
    colors['cyan'] = makeColor(curses.COLOR_CYAN, curses.COLOR_BLACK)
    colors['magenta'] = makeColor(curses.COLOR_MAGENTA, curses.COLOR_BLACK)

    colors['red_bg'] = makeColor(curses.COLOR_WHITE, curses.COLOR_RED)
    colors['blue_bg'] = makeColor(curses.COLOR_WHITE, curses.COLOR_BLUE)
    colors['green_bg'] = makeColor(curses.COLOR_BLACK, curses.COLOR_GREEN)
    colors['brown_bg'] = colors['yellow_bg'] = makeColor(curses.COLOR_BLACK, curses.COLOR_YELLOW)
    colors['cyan_bg'] = makeColor(curses.COLOR_BLACK, curses.COLOR_CYAN)
    colors['magenta_bg'] = makeColor(curses.COLOR_BLACK, curses.COLOR_MAGENTA)

    return f(stdscr, *args)

def vd():
    global g_vd
    if not g_vd:
        g_vd = VisiData()

    return g_vd

def status(s):
    return vd().status(s)

def run():
    os.putenv('ESCDELAY', '25')  # reduce ESC timeout to 25ms. http://en.chys.info/2009/09/esdelay-ncurses/
    ret = wrapper(curses_main)
    if ret:
        print(ret)

def curses_main(_scr):
    global scr
    scr = _scr

    # get control keys instead of signals
    curses.raw()

    # enable mouse events
#    curses.mousemask(curses.ALL_MOUSE_EVENTS)

    try:
        return vd().run()
    except Exception as e:
        if options.debug:
            raise
        return '%s: %s' % (type(e).__name__, str(e))


def wrapper(f, *args):
    import curses
    return curses.wrapper(setupcolors, f, *args)

setup_sheet_commands()  # on module import